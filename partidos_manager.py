import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule


def inicializar_lista_partidos():
    """Inicializa la lista de partidos en session_state."""
    if "lista_partidos" not in st.session_state:
        st.session_state.lista_partidos = []


def agregar_partido_a_lista(
    equipo_local: str,
    equipo_visitante: str,
    prob_tablas: dict,
) -> bool:
    """Agrega un partido a la lista usando SOLO las probabilidades de las tablas."""

    if equipo_local == equipo_visitante:
        st.error("❌ No puedes seleccionar el mismo equipo dos veces.")
        return False

    if not prob_tablas:
        st.error("❌ No se pudieron calcular las probabilidades de las tablas.")
        return False

    partido = {
        # Identificación mínima
        "Local": equipo_local,
        "Visitante": equipo_visitante,
        "Fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),

        # ===== Resultado y Dobles (Tabla 1) =====
        "Local Gana %": prob_tablas["Local_gana"],
        "Empate %": prob_tablas["Empate"],
        "Visitante Gana %": prob_tablas["Visitante_gana"],
        "1X %": prob_tablas["1X"],
        "X2 %": prob_tablas["X2"],
        "12 %": prob_tablas["12"],

        # ===== Goles en el Partido (Tabla 2) =====
        "Over 0.5 %": prob_tablas["Over_0.5_partido"],
        "Under 0.5 %": prob_tablas["Under_0.5_partido"],
        "Over 1.5 %": prob_tablas["Over_1.5_partido"],
        "Under 1.5 %": prob_tablas["Under_1.5_partido"],
        "Over 2.5 %": prob_tablas["Over_2.5_partido"],
        "Under 2.5 %": prob_tablas["Under_2.5_partido"],
        "Over 3.5 %": prob_tablas["Over_3.5_partido"],
        "Under 3.5 %": prob_tablas["Under_3.5_partido"],

        # ===== Goles por Equipo (Tabla 3) =====
        "Local Over 0.5 %": prob_tablas["Local_over_0.5"],
        "Local Under 0.5 %": prob_tablas["Local_under_0.5"],
        "Visitante Over 0.5 %": prob_tablas["Visitante_over_0.5"],
        "Visitante Under 0.5 %": prob_tablas["Visitante_under_0.5"],
        "Local Over 1.5 %": prob_tablas["Local_over_1.5"],
        "Local Under 1.5 %": prob_tablas["Local_under_1.5"],
        "Visitante Over 1.5 %": prob_tablas["Visitante_over_1.5"],
        "Visitante Under 1.5 %": prob_tablas["Visitante_under_1.5"],

        # ===== BTTS =====
        "BTTS %": prob_tablas["BTTS"],
        "NO BTTS %": prob_tablas["NO_BTTS"],

        # ===== Goles en el 1T (Tabla 4) =====
        "Over 0.5 1T %": prob_tablas["Over_0.5_1T"],
        "Under 0.5 1T %": prob_tablas["Under_0.5_1T"],
        "Over 1.5 1T %": prob_tablas["Over_1.5_1T"],
        "Under 1.5 1T %": prob_tablas["Under_1.5_1T"],
        "Local Over 0.5 1T %": prob_tablas["Local_over_0.5_1T"],
        "Local Under 0.5 1T %": prob_tablas["Local_under_0.5_1T"],
        "Visitante Over 0.5 1T %": prob_tablas["Visitante_over_0.5_1T"],
        "Visitante Under 0.5 1T %": prob_tablas["Visitante_under_0.5_1T"],
    }

    st.session_state.lista_partidos.append(partido)
    return True


def mostrar_tabla_partidos():
    if not st.session_state.lista_partidos:
        st.info("📭 No hay partidos en la lista.")
        return

    df_display = pd.DataFrame(st.session_state.lista_partidos)
    if "fecha" in df_display.columns:
        df_display = df_display.drop(columns="Fecha")
    st.dataframe(df_display, use_container_width=True)


def crear_excel_descargable():
    if not st.session_state.lista_partidos:
        return None

    df_export = pd.DataFrame(st.session_state.lista_partidos)
    if "Fecha" in df_export.columns:
        df_export = df_export.drop(columns=["Fecha"])
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Partidos", index=False)

        workbook = writer.book
        ws = writer.sheets["Partidos"]

        # ===== ESTILOS BASE =====
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Aplicar estilos al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Ajustar ancho de columnas
        for idx, col in enumerate(df_export.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = 16

        # ===== FORMATO CONDICIONAL =====
        # Definir colores
        amarillo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        azul_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        verde_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

        # Colores de texto para mejor contraste
        blanco_font = Font(color="FFFFFF", bold=True)
        negro_font = Font(color="000000", bold=True)

        # Columnas a las que aplicar formato condicional (todas las que terminan en %)
        columnas_porcentaje = [
            col for col in df_export.columns if "%" in col
        ]

        # Aplicar formato a cada fila (empezar desde fila 2, porque fila 1 es encabezado)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                col_letter = cell.column_letter
                col_name = df_export.columns[cell.column - 1]

                # Solo aplicar formato si la columna es porcentaje
                if col_name in columnas_porcentaje:
                    cell_ref = f"{col_letter}{row_idx}"
                    
                    # AMARILLO: 50-58
                    rule_amarillo = CellIsRule(
                        operator="between",
                        formula=["50", "58"],
                        fill=amarillo_fill,
                        font=negro_font,
                    )
                    ws.conditional_formatting.add(cell_ref, rule_amarillo)

                    # AZUL: 58-75
                    rule_azul = CellIsRule(
                        operator="between",
                        formula=["58", "75"],
                        fill=azul_fill,
                        font=blanco_font,
                    )
                    ws.conditional_formatting.add(cell_ref, rule_azul)

                    # VERDE: >= 75
                    rule_verde = CellIsRule(
                        operator="greaterThanOrEqual",
                        formula=["75"],
                        fill=verde_fill,
                        font=blanco_font,
                    )
                    ws.conditional_formatting.add(cell_ref, rule_verde)

                # Formato de número (2 decimales)
                if col_name in columnas_porcentaje:
                    cell.number_format = "0.0"

                # Alineación para todas las celdas
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    output.seek(0)
    return output


def mostrar_boton_agregar_partido(
    equipo_local: str,
    equipo_visitante: str,
    prob_tablas: dict,
):
    st.markdown("---")
    st.markdown("### 🗂️ Lista de Partidos Guardados")

    if st.session_state.lista_partidos:
        mostrar_tabla_partidos()

        col_limp, col_dl = st.columns(2)

        with col_limp:
            if st.button("🔄 Limpiar Lista", use_container_width=True):
                st.session_state.lista_partidos = []
                st.success("✅ Lista vaciada correctamente")

        with col_dl:
            excel_data = crear_excel_descargable()
            if excel_data:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    "📥 Descargar XLSX con Formato",
                    data=excel_data,
                    file_name=f"partidos_guardados_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    else:
        st.info("📭 No hay partidos en la lista. Agrega uno con el botón de arriba.")